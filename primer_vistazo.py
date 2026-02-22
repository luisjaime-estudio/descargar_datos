"""
Analizador de calidad de archivos NetCDF descargados.

Realiza un analisis exhaustivo de archivos NetCDF descargados,
verificando la integridad de los datos, completitud de las descargas, y
generando metricas de calidad tanto a nivel de archivo como de datos.

Salidas generadas:
    - metricas_calidad.csv    : metricas detalladas por anio de inicializacion
    - informe_calidad_datos.xlsx : reporte Excel con 3 pestanas
"""

import re
from pathlib import Path
from typing import Dict, List, Optional, Any

import netCDF4 as nc
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
VARIABLES_COORDENADAS: set = {
    'time', 'lat', 'lon', 'latitude', 'longitude',
    'bnds', 'time_bnds', 'lat_bnds', 'lon_bnds',
}

PATRON_NOMBRE_ARCHIVO: str = (
    r"^(\w+)_(\w+)_([^_]+)_([^_]+)_([^_]+)_([^_]+)_([^_]+)\.nc$"
)

SEPARADOR: str = "=" * 80
SUB_SEPARADOR: str = "-" * 80


class AnalizadorNetCDF:
    """Analizador de archivos NetCDF para verificacion de calidad.

    Atributos de estado (se pueblan durante la ejecucion):
        directorio          : Path del directorio raiz
        archivos_nc         : lista de archivos .nc encontrados
        df_archivos         : DataFrame con metadatos de archivos
        completitud         : dict con info de completitud por combinacion
        df_metricas_datos   : DataFrame con metricas de cada archivo
        df_metricas_resumen : DataFrame con metricas por anio/combinacion
    """

    # =====================================================================
    # Inicializacion
    # =====================================================================
    def __init__(self, directorio: str = ".") -> None:
        """Inicializa el analizador y descubre archivos .nc recursivamente."""
        self.directorio: Path = Path(directorio)
        self.archivos_nc: List[Path] = list(self.directorio.rglob('*.nc'))

        # Estado que se puebla durante el analisis
        self.df_archivos: pd.DataFrame = pd.DataFrame()
        self.completitud: Dict[str, Dict] = {}
        self.df_metricas_datos: pd.DataFrame = pd.DataFrame()
        self.df_metricas_resumen: pd.DataFrame = pd.DataFrame()

    # =====================================================================
    # ORQUESTADOR PRINCIPAL
    # =====================================================================
    def ejecutar_analisis(self) -> None:
        """Orquestador: ejecuta la pipeline completa de analisis.

        Cada paso delega en metodos auxiliares privados.
        El flujo se lee como un indice del analisis.
        """
        self._imprimir_cabecera_analisis()

        if not self.archivos_nc:
            print("\n[X] No se encontraron archivos .nc en el directorio")
            return

        # 1. Metadatos de archivos
        self._analizar_metadatos()

        # 2. Verificacion de completitud
        self._verificar_completitud()

        # 3. Anomalias de tamano
        self._analizar_anomalias_tamano()

        # 4. Contenido NetCDF (muestra)
        self._analizar_contenido_muestra(tamano_muestra=2)

        # 5. Calidad de datos (muestra)
        self._analizar_calidad_muestra(tamano_muestra=3)

        # 6. Metricas por archivo
        self._calcular_metricas_por_archivo()

        # 7. CSV de metricas
        self._generar_csv_metricas()

        # 8. Reporte Excel
        self._generar_reporte_excel()

        # 9. Resumen en consola
        self._imprimir_resumen_consola()

        # 10. Mensaje final
        self._imprimir_mensaje_final()

    # =====================================================================
    # Utilidades de parseo de nombres
    # =====================================================================
    def _parsear_nombre_archivo(self, nombre: str) -> Dict[str, Any]:
        """Extrae componentes CMIP6 del nombre de un archivo .nc.

        Formato: variable_table_source_experiment_member_grid_timerange.nc
        """
        try:
            coincidencia = re.match(PATRON_NOMBRE_ARCHIVO, nombre)
            if coincidencia:
                return {
                    "variable_id":   coincidencia.group(1),
                    "table_id":      coincidencia.group(2),
                    "source_id":     coincidencia.group(3),
                    "experiment_id": coincidencia.group(4),
                    "member_id":     coincidencia.group(5),
                    "grid_label":    coincidencia.group(6),
                    "time_range":    coincidencia.group(7),
                    "filename":      nombre,
                }
            return {"filename": nombre, "parse_error": True}
        except Exception as e:
            print(f"  [!] Error parseando nombre '{nombre}': {e}")
            return {"filename": nombre, "parse_error": True}

    @staticmethod
    def _extraer_anio_inicializacion(member_id: str) -> Optional[int]:
        """Extrae el anio de inicializacion (ej: 1961 de 's1961-r1i1p1f1')."""
        try:
            coincidencia = re.search(r's(\d{4})', member_id)
            return int(coincidencia.group(1)) if coincidencia else None
        except Exception:
            return None

    @staticmethod
    def _extraer_variant_label(member_id: str) -> Optional[str]:
        """Extrae el variant_label (ej: 'r1i1p1f1' de 's1961-r1i1p1f1')."""
        try:
            coincidencia = re.search(r'-(r\d+i\d+p\d+f\d+)', member_id)
            return coincidencia.group(1) if coincidencia else None
        except Exception:
            return None

    # =====================================================================
    # Utilidades de datos
    # =====================================================================
    @staticmethod
    def _obtener_variables_datos(dataset: nc.Dataset) -> List[str]:
        """Devuelve los nombres de variables que NO son coordenadas."""
        return [v for v in dataset.variables.keys()
                if v not in VARIABLES_COORDENADAS]

    @staticmethod
    def _obtener_datos_validos(variable: Any, datos: np.ndarray
                               ) -> tuple:
        """Calcula datos validos y porcentaje de faltantes.

        Returns:
            (datos_validos, porcentaje_faltantes)
        """
        if hasattr(datos, 'mask'):
            validos = datos[~datos.mask]
            pct_faltantes = (datos.mask.sum() / datos.size) * 100
        elif np.issubdtype(datos.dtype, np.floating):
            mascara_nan = np.isnan(datos)
            validos = datos[~mascara_nan]
            pct_faltantes = (mascara_nan.sum() / datos.size) * 100
        else:
            validos = datos.ravel()
            pct_faltantes = 0.0

        return validos, pct_faltantes

    # =====================================================================
    # PASO 1: Analisis de metadatos
    # =====================================================================
    def _analizar_metadatos(self) -> None:
        """Parsea todos los archivos .nc y construye self.df_archivos."""
        try:
            self._imprimir_seccion("ANALISIS DE METADATOS DE ARCHIVOS")

            info_archivos: List[Dict] = []
            for archivo_nc in tqdm(self.archivos_nc, desc="Analizando metadatos"):
                info = self._extraer_info_archivo(archivo_nc)
                info_archivos.append(info)

            self.df_archivos = pd.DataFrame(info_archivos)
            self._imprimir_estadisticas_tamano()

        except Exception as e:
            print(f"  [X] Error en analisis de metadatos: {e}")

    def _extraer_info_archivo(self, archivo_nc: Path) -> Dict[str, Any]:
        """Extrae metadatos de un unico archivo .nc."""
        info = self._parsear_nombre_archivo(archivo_nc.name)
        info["file_size_mb"] = archivo_nc.stat().st_size / (1024 * 1024)
        info["file_path"] = str(archivo_nc)

        if "member_id" in info:
            info["init_year"] = self._extraer_anio_inicializacion(info["member_id"])
            info["variant_label"] = self._extraer_variant_label(info["member_id"])

        return info

    def _imprimir_estadisticas_tamano(self) -> None:
        """Imprime estadisticas basicas de tamano del conjunto."""
        df = self.df_archivos
        print(f"\nTotal de archivos encontrados: {len(df)}")
        print(f"Tamano total: {df['file_size_mb'].sum():.2f} MB")
        print(f"\nTamano promedio por archivo: {df['file_size_mb'].mean():.2f} MB")
        print(f"Tamano minimo: {df['file_size_mb'].min():.2f} MB")
        print(f"Tamano maximo: {df['file_size_mb'].max():.2f} MB")

    # =====================================================================
    # PASO 2: Verificacion de completitud
    # =====================================================================
    def _verificar_completitud(self) -> None:
        """Verifica completitud temporal y puebla self.completitud."""
        try:
            self._imprimir_seccion("VERIFICACION DE COMPLETITUD")

            df = self._filtrar_archivos_parseados()
            combinaciones = self._obtener_combinaciones(df)

            print(f"\nCombinaciones de source_id y variant_label encontradas:")
            print(combinaciones.to_string(index=False))

            self._imprimir_sub_seccion("ANIOS DE INICIALIZACION POR COMBINACION")

            for _, fila in combinaciones.iterrows():
                self._detectar_anios_faltantes(df, fila['source_id'],
                                               fila['variant_label'])

        except Exception as e:
            print(f"  [X] Error en verificacion de completitud: {e}")

    def _filtrar_archivos_parseados(self) -> pd.DataFrame:
        """Filtra archivos que no pudieron ser parseados. Devuelve df limpio."""
        df = self.df_archivos.copy()
        if 'parse_error' in df.columns and df['parse_error'].any():
            print("\n[!]  ADVERTENCIA: Algunos archivos no pudieron ser parseados")
            df = df[~df['parse_error'].fillna(False)]
        return df

    @staticmethod
    def _obtener_combinaciones(df: pd.DataFrame) -> pd.DataFrame:
        """Agrupa y cuenta por source_id / variant_label."""
        return (df.groupby(['source_id', 'variant_label'])
                  .size()
                  .reset_index(name='count'))

    def _detectar_anios_faltantes(self, df: pd.DataFrame,
                                   source: str, variant: str) -> None:
        """Analiza un par source/variant y registra resultado en self.completitud."""
        subconjunto = df[(df['source_id'] == source) &
                         (df['variant_label'] == variant)]
        anios = [int(x) for x in sorted(subconjunto['init_year'].dropna().unique())]

        print(f"\nSource: {source} | Variant: {variant}")
        print(f"  Archivos: {len(subconjunto)}")
        print(f"  Anios de inicializacion: {len(anios)}")
        print(f"  Rango: {min(anios)} - {max(anios)}")

        esperados = set(range(min(anios), max(anios) + 1))
        faltantes = sorted(esperados - set(anios))

        if faltantes:
            print(f"  [!]  Anios faltantes: {faltantes}")
        else:
            print(f"  [OK] Todos los anios presentes en el rango")

        self.completitud[f"{source}_{variant}"] = {
            "total_files":    len(subconjunto),
            "init_years":     anios,
            "missing_years":  faltantes,
            "expected_size_mb": subconjunto['file_size_mb'].mean(),
            "total_size_mb":  subconjunto['file_size_mb'].sum(),
        }

    # =====================================================================
    # PASO 3: Anomalias de tamano
    # =====================================================================
    def _analizar_anomalias_tamano(self, umbral: float = 0.05) -> None:
        """Detecta archivos con tamano anomalo por combinacion."""
        try:
            self._imprimir_seccion("ANALISIS DE ANOMALIAS EN TAMANO DE ARCHIVO")
            df = self._filtrar_archivos_parseados()

            for (source, variant), grupo in df.groupby(['source_id', 'variant_label']):
                self._reportar_anomalia_grupo(source, variant, grupo, umbral)

        except Exception as e:
            print(f"  [X] Error en analisis de anomalias: {e}")

    @staticmethod
    def _reportar_anomalia_grupo(source: str, variant: str,
                                  grupo: pd.DataFrame, umbral: float) -> None:
        """Imprime resultado de anomalias para un grupo source/variant."""
        media = grupo['file_size_mb'].mean()
        desv = grupo['file_size_mb'].std()

        print(f"\nSource: {source} | Variant: {variant}")
        print(f"  Tamano esperado: {media:.2f} +/- {desv:.2f} MB")

        anomalos = grupo[
            (grupo['file_size_mb'] < media * (1 - umbral)) |
            (grupo['file_size_mb'] > media * (1 + umbral))
        ]

        if len(anomalos) > 0:
            print(f"  [!]  {len(anomalos)} archivo(s) con tamano anomalo:")
            for _, fila in anomalos.iterrows():
                desviacion = ((fila['file_size_mb'] - media) / media) * 100
                print(f"    - {fila['filename']}: {fila['file_size_mb']:.2f} MB "
                      f"({desviacion:+.1f}%)")
        else:
            print(f"  [OK] Todos los archivos tienen tamano consistente")

    # =====================================================================
    # PASO 4: Contenido NetCDF (muestra)
    # =====================================================================
    def _analizar_contenido_muestra(self, tamano_muestra: int = 3) -> None:
        """Inspecciona dimensiones, variables y atributos de una muestra."""
        try:
            self._imprimir_seccion("ANALISIS DE CONTENIDO NetCDF")

            for archivo_nc in self.archivos_nc[:tamano_muestra]:
                self._inspeccionar_archivo_nc(archivo_nc)

        except Exception as e:
            print(f"  [X] Error en analisis de contenido: {e}")

    def _inspeccionar_archivo_nc(self, archivo_nc: Path) -> None:
        """Inspecciona un unico archivo NetCDF."""
        print(f"\n{SEPARADOR}")
        print(f"Archivo: {archivo_nc.name}")
        print(f"{SEPARADOR}")

        try:
            with nc.Dataset(archivo_nc, 'r') as dataset:
                self._imprimir_dimensiones(dataset)
                self._imprimir_variables(dataset)
                self._imprimir_atributos_globales(dataset)
        except Exception as e:
            print(f"  [X] Error al leer archivo: {e}")

    @staticmethod
    def _imprimir_dimensiones(dataset: nc.Dataset) -> None:
        """Imprime las dimensiones de un dataset NetCDF."""
        print("\n--- DIMENSIONES ---")
        for nombre, dim in dataset.dimensions.items():
            ilimitada = '(unlimited)' if dim.isunlimited() else ''
            print(f"  {nombre}: {len(dim)} {ilimitada}")

    @staticmethod
    def _imprimir_variables(dataset: nc.Dataset) -> None:
        """Imprime informacion de las variables de un dataset NetCDF."""
        print("\n--- VARIABLES ---")
        for nombre, var in dataset.variables.items():
            print(f"  {nombre}:")
            print(f"    Dimensiones: {var.dimensions}")
            print(f"    Shape: {var.shape}")
            print(f"    Tipo: {var.dtype}")
            if hasattr(var, 'units'):
                print(f"    Unidades: {var.units}")
            if hasattr(var, 'long_name'):
                print(f"    Nombre: {var.long_name}")

    @staticmethod
    def _imprimir_atributos_globales(dataset: nc.Dataset) -> None:
        """Imprime los atributos globales de un dataset NetCDF."""
        print("\n--- ATRIBUTOS GLOBALES ---")
        for nombre_attr in dataset.ncattrs():
            print(f"  {nombre_attr}: {dataset.getncattr(nombre_attr)}")

    # =====================================================================
    # PASO 5: Calidad de datos (muestra)
    # =====================================================================
    def _analizar_calidad_muestra(self, tamano_muestra: int = 5) -> None:
        """Calcula estadisticas de calidad sobre una muestra de archivos."""
        try:
            self._imprimir_seccion("ANALISIS DE CALIDAD DE DATOS")

            for archivo_nc in self.archivos_nc[:tamano_muestra]:
                self._evaluar_calidad_archivo(archivo_nc)

        except Exception as e:
            print(f"  [X] Error en analisis de calidad: {e}")

    def _evaluar_calidad_archivo(self, archivo_nc: Path) -> None:
        """Evalua la calidad de datos de un unico archivo."""
        print(f"\n{archivo_nc.name}")
        print(SUB_SEPARADOR)

        try:
            with nc.Dataset(archivo_nc, 'r') as dataset:
                vars_datos = self._obtener_variables_datos(dataset)
                for nombre_var in vars_datos[:1]:
                    self._imprimir_estadisticas_variable(dataset, nombre_var)
                    self._imprimir_info_temporal(dataset)
        except Exception as e:
            print(f"  [X] Error al analizar datos: {e}")

    def _imprimir_estadisticas_variable(self, dataset: nc.Dataset,
                                         nombre_var: str) -> None:
        """Imprime estadisticas de una variable de datos."""
        var = dataset.variables[nombre_var]
        datos = var[:]
        validos, pct_faltantes = self._obtener_datos_validos(var, datos)

        print(f"\nVariable: {nombre_var}")
        print(f"  Shape: {datos.shape}")
        print(f"  Tipo de dato: {datos.dtype}")

        if hasattr(datos, 'mask'):
            print(f"  Valores faltantes: {datos.mask.sum()} ({pct_faltantes:.2f}%)")
        else:
            n_faltantes = int(round(pct_faltantes * datos.size / 100))
            print(f"  Valores faltantes (NaN): {n_faltantes} ({pct_faltantes:.2f}%)")

        if len(validos) > 0:
            print(f"  Minimo: {validos.min():.6f}")
            print(f"  Maximo: {validos.max():.6f}")
            print(f"  Media: {validos.mean():.6f}")
            print(f"  Desviacion estandar: {validos.std():.6f}")

            if hasattr(var, '_FillValue'):
                print(f"  _FillValue: {var._FillValue}")

            if np.issubdtype(datos.dtype, np.floating):
                n_inf = np.isinf(datos).sum()
                if n_inf > 0:
                    print(f"  [!]  Valores infinitos: {n_inf}")

    @staticmethod
    def _imprimir_info_temporal(dataset: nc.Dataset) -> None:
        """Imprime informacion de la dimension temporal si existe."""
        if 'time' not in dataset.variables:
            return
        var_tiempo = dataset.variables['time']
        print(f"\n  Dimension temporal:")
        print(f"    Pasos de tiempo: {len(var_tiempo)}")
        if hasattr(var_tiempo, 'units'):
            print(f"    Unidades: {var_tiempo.units}")
        if hasattr(var_tiempo, 'calendar'):
            print(f"    Calendario: {var_tiempo.calendar}")

    # =====================================================================
    # PASO 6: Metricas por archivo
    # =====================================================================
    def _calcular_metricas_por_archivo(self) -> None:
        """Calcula metricas de calidad para cada archivo. Puebla self.df_metricas_datos."""
        try:
            self._imprimir_seccion("CALCULO DE METRICAS DE DATOS")

            lista_metricas: List[Dict] = []
            for _, fila in tqdm(self.df_archivos.iterrows(),
                                total=len(self.df_archivos),
                                desc="Calculando metricas"):
                metrica = self._extraer_metricas_archivo(fila)
                if metrica is not None:
                    lista_metricas.append(metrica)

            self.df_metricas_datos = pd.DataFrame(lista_metricas)

        except Exception as e:
            print(f"  [X] Error calculando metricas: {e}")

    def _extraer_metricas_archivo(self, fila: pd.Series) -> Optional[Dict]:
        """Extrae metricas de un unico archivo a partir de su fila de metadatos."""
        archivo_nc = Path(fila['file_path'])

        try:
            with nc.Dataset(archivo_nc, 'r') as dataset:
                vars_datos = self._obtener_variables_datos(dataset)
                if not vars_datos:
                    return None

                nombre_var = vars_datos[0]
                var = dataset.variables[nombre_var]
                datos = var[:]
                validos, pct_faltantes = self._obtener_datos_validos(var, datos)

                return {
                    'filename':      fila['filename'],
                    'source_id':     fila.get('source_id'),
                    'variant_label': fila.get('variant_label'),
                    'init_year':     fila.get('init_year'),
                    'variable':      nombre_var,
                    'missing_pct':   pct_faltantes,
                    'min_value':     float(validos.min()) if len(validos) > 0 else np.nan,
                    'max_value':     float(validos.max()) if len(validos) > 0 else np.nan,
                    'mean_value':    float(validos.mean()) if len(validos) > 0 else np.nan,
                    'std_value':     float(validos.std()) if len(validos) > 0 else np.nan,
                    'time_steps':    (len(dataset.variables['time'])
                                      if 'time' in dataset.variables else 0),
                }

        except Exception as e:
            print(f"  [!]  Error procesando {fila['filename']}: {e}")
            return None

    # =====================================================================
    # PASO 7: CSV de metricas
    # =====================================================================
    def _generar_csv_metricas(self) -> None:
        """Genera CSV con metricas por anio/combinacion. Puebla self.df_metricas_resumen."""
        try:
            self._imprimir_seccion("GENERACION DE CSV DE METRICAS")

            filas: List[Dict] = []
            for nombre_combo, info in self.completitud.items():
                source_id, variant_label = nombre_combo.split('_', 1)
                for anio in info['init_years']:
                    fila = self._construir_fila_metrica(
                        source_id, variant_label, anio)
                    if fila is not None:
                        filas.append(fila)

            self.df_metricas_resumen = (
                pd.DataFrame(filas)
                  .sort_values(['source_id', 'variant_label', 'anio_inicializacion'])
            )

            ruta_salida = self.directorio / "metricas_calidad.csv"
            self.df_metricas_resumen.to_csv(ruta_salida, index=False)

            print(f"\n[DONE] Metricas guardadas en: {ruta_salida}")
            print(f"\nColumnas incluidas:")
            for col in self.df_metricas_resumen.columns:
                print(f"  - {col}")
            print(f"\nTotal de filas: {len(self.df_metricas_resumen)}")

        except Exception as e:
            print(f"  [X] Error generando CSV de metricas: {e}")

    def _construir_fila_metrica(self, source_id: str, variant_label: str,
                                 anio: int) -> Optional[Dict]:
        """Construye una fila del CSV de metricas para un anio concreto."""
        df = self.df_archivos
        dm = self.df_metricas_datos

        archivo = df[(df['source_id'] == source_id) &
                      (df['variant_label'] == variant_label) &
                      (df['init_year'] == anio)]
        datos = dm[(dm['source_id'] == source_id) &
                    (dm['variant_label'] == variant_label) &
                    (dm['init_year'] == anio)]

        if len(archivo) == 0:
            return None

        tiene_datos = len(datos) > 0
        return {
            'source_id':                   source_id,
            'variant_label':               variant_label,
            'anio_inicializacion':         int(anio),
            'tamano_archivo_mb':           archivo['file_size_mb'].iloc[0],
            'porcentaje_valores_faltantes': datos['missing_pct'].iloc[0] if tiene_datos else np.nan,
            'valor_dato_min':              datos['min_value'].iloc[0]   if tiene_datos else np.nan,
            'valor_dato_max':              datos['max_value'].iloc[0]   if tiene_datos else np.nan,
            'valor_dato_media':            datos['mean_value'].iloc[0]  if tiene_datos else np.nan,
            'valor_dato_std':              datos['std_value'].iloc[0]   if tiene_datos else np.nan,
            'pasos_temporales':            int(datos['time_steps'].iloc[0]) if tiene_datos else np.nan,
        }

    # =====================================================================
    # PASO 8: Reporte Excel
    # =====================================================================
    def _generar_reporte_excel(self) -> None:
        """Genera reporte Excel con 3 pestanas."""
        try:
            self._imprimir_seccion("GENERACION DE REPORTE EXCEL")

            ruta_salida = self.directorio / "informe_calidad_datos_v2.xlsx"
            estilos = self._crear_estilos_excel()

            wb = Workbook()
            wb.remove(wb.active)

            self._escribir_pestana_resumen(wb, estilos)
            # self._escribir_pestana_detalle(wb, estilos)
            # self._escribir_pestana_agregados(wb, estilos)

            wb.save(ruta_salida)

            print(f"\n[DONE] Reporte Excel guardado en: {ruta_salida}")
            print(f"\nPestanas incluidas:")
            print(f"  1. Resumen Ejecutivo - Informe general de calidad")
            # print(f"  2. Metricas Detalladas - Datos por anio de inicializacion")
            # print(f"  3. Metricas Agregadas - Resumen por combinacion source_id/variant_label")

        except Exception as e:
            print(f"  [X] Error generando reporte Excel: {e}")

    @staticmethod
    def _crear_estilos_excel() -> Dict[str, Any]:
        """Crea y devuelve los estilos reutilizables para el Excel."""
        return {
            'header_fill': PatternFill(start_color="366092",
                                       end_color="366092",
                                       fill_type="solid"),
            'header_font': Font(bold=True, color="FFFFFF"),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin'),
            ),
        }

    def _escribir_pestana_resumen(self, wb: Workbook,
                                   estilos: Dict) -> None:
        """Escribe la pestana 'Resumen Ejecutivo'."""
        ws = wb.create_sheet("Resumen Ejecutivo")

        # Titulo
        ws['A1'] = "INFORME DE CALIDAD DE DATOS NetCDF"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        fila = 3
        fila = self._escribir_estadisticas_generales(ws, fila)
        fila = self._escribir_cobertura_temporal(ws, fila)
        fila = self._escribir_calidad_datos_excel(ws, fila)
        self._escribir_detalles_combinaciones(ws, fila)

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

    def _escribir_estadisticas_generales(self, ws: Any, fila: int) -> int:
        """Bloque 'Estadisticas Generales' en la pestana resumen."""
        ws[f'A{fila}'] = "ESTADISTICAS GENERALES"
        ws[f'A{fila}'].font = Font(bold=True, size=12)
        fila += 1

        ws[f'A{fila}'] = "Total de archivos:"
        ws[f'B{fila}'] = len(self.df_archivos)
        fila += 1

        ws[f'A{fila}'] = "Tamano total (MB):"
        ws[f'B{fila}'] = round(self.df_archivos['file_size_mb'].sum(), 2)
        fila += 1

        ws[f'A{fila}'] = "Combinaciones source_id/variant_label:"
        ws[f'B{fila}'] = len(self.completitud)
        fila += 2

        return fila

    def _escribir_cobertura_temporal(self, ws: Any, fila: int) -> int:
        """Bloque 'Cobertura Temporal' en la pestana resumen."""
        ws[f'A{fila}'] = "COBERTURA TEMPORAL"
        ws[f'A{fila}'].font = Font(bold=True, size=12)
        fila += 1

        if 'init_year' in self.df_archivos.columns:
            anios = self.df_archivos['init_year'].dropna()
            if len(anios) > 0:
                ws[f'A{fila}'] = "Anios de inicializacion:"
                ws[f'B{fila}'] = f"{int(anios.min())} - {int(anios.max())}"
                fila += 1

                ws[f'A{fila}'] = "Total de anios unicos:"
                ws[f'B{fila}'] = len(anios.unique())
                fila += 2

        return fila

    def _escribir_calidad_datos_excel(self, ws: Any, fila: int) -> int:
        """Bloque 'Calidad de Datos' en la pestana resumen."""
        ws[f'A{fila}'] = "CALIDAD DE DATOS"
        ws[f'A{fila}'].font = Font(bold=True, size=12)
        fila += 1

        total_faltantes = sum(len(v['missing_years'])
                              for v in self.completitud.values())
        ws[f'A{fila}'] = "Anios faltantes:"
        ws[f'B{fila}'] = total_faltantes

        if total_faltantes == 0:
            ws[f'C{fila}'] = "[OK] Completo"
            ws[f'C{fila}'].font = Font(color="00B050")
        else:
            ws[f'C{fila}'] = "[!] Incompleto"
            ws[f'C{fila}'].font = Font(color="FF0000")
        fila += 2

        return fila

    def _escribir_detalles_combinaciones(self, ws: Any, fila: int) -> None:
        """Bloque 'Detalles por Combinacion' en la pestana resumen."""
        ws[f'A{fila}'] = "DETALLES POR COMBINACION"
        ws[f'A{fila}'].font = Font(bold=True, size=12)
        fila += 1

        for nombre_combo, info in self.completitud.items():
            ws[f'A{fila}'] = nombre_combo
            ws[f'A{fila}'].font = Font(bold=True)
            fila += 1

            ws[f'A{fila}'] = "  Archivos:"
            ws[f'B{fila}'] = info['total_files']
            fila += 1

            ws[f'A{fila}'] = "  Tamano esperado (MB):"
            ws[f'B{fila}'] = round(info['expected_size_mb'], 2)
            fila += 1

            if info['missing_years']:
                ws[f'A{fila}'] = "  Anios faltantes:"
                ws[f'B{fila}'] = ', '.join(map(str, info['missing_years']))
                fila += 1

            fila += 1

    def _escribir_pestana_detalle(self, wb: Workbook,
                                   estilos: Dict) -> None:
        """Escribe la pestana 'Metricas Detalladas'."""
        ws = wb.create_sheet("Metricas Detalladas")
        self._volcar_dataframe_a_hoja(ws, self.df_metricas_resumen, estilos)

    def _escribir_pestana_agregados(self, wb: Workbook,
                                     estilos: Dict) -> None:
        """Escribe la pestana 'Metricas Agregadas'."""
        ws = wb.create_sheet("Metricas Agregadas")
        df_agregado = self._generar_metricas_agregadas()
        self._volcar_dataframe_a_hoja(ws, df_agregado, estilos)

    def _volcar_dataframe_a_hoja(self, ws: Any, df: pd.DataFrame,
                                  estilos: Dict) -> None:
        """Vuelca un DataFrame a una hoja Excel con estilos y ajuste de anchos."""
        for r_idx, fila in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, valor in enumerate(fila, 1):
                celda = ws.cell(row=r_idx, column=c_idx, value=valor)
                celda.border = estilos['border']

                if r_idx == 1:
                    celda.fill = estilos['header_fill']
                    celda.font = estilos['header_font']
                    celda.alignment = Alignment(horizontal='center',
                                                vertical='center')

        self._ajustar_anchos_columnas(ws)

    @staticmethod
    def _ajustar_anchos_columnas(ws: Any) -> None:
        """Ajusta automaticamente los anchos de columna de una hoja."""
        for columna in ws.columns:
            ancho_max = 0
            letra = columna[0].column_letter
            for celda in columna:
                try:
                    largo = len(str(celda.value))
                    if largo > ancho_max:
                        ancho_max = largo
                except Exception:
                    pass
            ws.column_dimensions[letra].width = min(ancho_max + 2, 50)

    def _generar_metricas_agregadas(self) -> pd.DataFrame:
        """Genera metricas agregadas por combinacion source_id/variant_label."""
        metricas: List[Dict] = []

        for nombre_combo, info in self.completitud.items():
            source_id, variant_label = nombre_combo.split('_', 1)
            subconjunto = self.df_archivos[
                (self.df_archivos['source_id'] == source_id) &
                (self.df_archivos['variant_label'] == variant_label)
            ]

            metricas.append({
                'source_id':       source_id,
                'variant_label':   variant_label,
                'anio_init_min':   min(info['init_years']) if info['init_years'] else np.nan,
                'anio_init_max':   max(info['init_years']) if info['init_years'] else np.nan,
                'numero_archivos': info['total_files'],
                'tamano_medio_mb': subconjunto['file_size_mb'].mean(),
                'desv_est_tamano_mb': subconjunto['file_size_mb'].std(),
            })

        return pd.DataFrame(metricas)

    # =====================================================================
    # PASO 9: Resumen en consola
    # =====================================================================
    def _imprimir_resumen_consola(self) -> None:
        """Imprime resumen ejecutivo en consola."""
        try:
            self._imprimir_seccion("RESUMEN EJECUTIVO")

            self._resumen_estadisticas_generales()
            self._resumen_cobertura_temporal()
            self._resumen_calidad_datos()
            self._resumen_tamanos_esperados()

            print(f"\n{SEPARADOR}")

        except Exception as e:
            print(f"  [X] Error imprimiendo resumen: {e}")

    def _resumen_estadisticas_generales(self) -> None:
        """Bloque de estadisticas generales para consola."""
        print(f"\n[STATS] ESTADISTICAS GENERALES")
        print(f"  Total de archivos: {len(self.df_archivos)}")
        print(f"  Tamano total: {self.df_archivos['file_size_mb'].sum():.2f} MB")
        print(f"  Combinaciones source_id/variant_label: {len(self.completitud)}")

    def _resumen_cobertura_temporal(self) -> None:
        """Bloque de cobertura temporal para consola."""
        print(f"\n[DATE] COBERTURA TEMPORAL")
        if 'init_year' in self.df_archivos.columns:
            anios = self.df_archivos['init_year'].dropna()
            if len(anios) > 0:
                print(f"  Anios de inicializacion: {int(anios.min())} - {int(anios.max())}")
                print(f"  Total de anios unicos: {len(anios.unique())}")

    def _resumen_calidad_datos(self) -> None:
        """Bloque de calidad de datos para consola."""
        print(f"\n[CHECK] CALIDAD DE DATOS")
        total = sum(len(v['missing_years']) for v in self.completitud.values())
        if total == 0:
            print(f"  [OK] No se detectaron anios faltantes")
        else:
            print(f"  [!]  Se detectaron {total} anios faltantes en total")

    def _resumen_tamanos_esperados(self) -> None:
        """Bloque de tamanos esperados por combinacion para consola."""
        print(f"\n[SIZE] TAMANO ESPERADO POR COMBINACION")
        for nombre, info in self.completitud.items():
            print(f"  {nombre}: ~{info['expected_size_mb']:.2f} MB por archivo")

    # =====================================================================
    # PASO 10: Mensaje final
    # =====================================================================
    def _imprimir_mensaje_final(self) -> None:
        """Imprime mensaje final con archivos generados."""
        print(f"\n{SEPARADOR}")
        print("[DONE] ANALISIS COMPLETADO")
        print(SEPARADOR)
        print(f"\nArchivos generados:")
        print(f"  - metricas_calidad.csv          : Metricas detalladas por anio")
        print(f"  - informe_calidad_datos.xlsx     : Reporte Excel completo")
        print(f"\n{SEPARADOR}")

    # =====================================================================
    # Helpers de impresion
    # =====================================================================
    def _imprimir_cabecera_analisis(self) -> None:
        """Imprime la cabecera del analisis."""
        print(f"\n{SEPARADOR}")
        print("ANALISIS DE CALIDAD DE ARCHIVOS NetCDF")
        print(SEPARADOR)
        print(f"Directorio: {self.directorio.absolute()}")
        print(f"Archivos .nc encontrados: {len(self.archivos_nc)}")

    @staticmethod
    def _imprimir_seccion(titulo: str) -> None:
        """Imprime un separador de seccion."""
        print(f"\n{SEPARADOR}")
        print(titulo)
        print(SEPARADOR)

    @staticmethod
    def _imprimir_sub_seccion(titulo: str) -> None:
        """Imprime un sub-separador de seccion."""
        print(f"\n{SUB_SEPARADOR}")
        print(titulo)
        print(SUB_SEPARADOR)


# ===========================================================================
# Punto de entrada
# ===========================================================================
def main() -> None:
    """Funcion principal."""
    analizador = AnalizadorNetCDF(directorio=".")
    analizador.ejecutar_analisis()


if __name__ == "__main__":
    main()
