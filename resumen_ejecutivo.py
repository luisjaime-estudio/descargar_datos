"""
Genera la hoja 'Resumen Ejecutivo' del informe de calidad NetCDF.
"""

import re
from pathlib import Path
from typing import Dict, List, Optional, Any

import netCDF4 as nc
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
VARIABLES_COORDENADAS = {
    'time', 'lat', 'lon', 'latitude', 'longitude',
    'bnds', 'time_bnds', 'lat_bnds', 'lon_bnds',
}
PATRON_NOMBRE = r"^(\w+)_(\w+)_([^_]+)_([^_]+)_([^_]+)_([^_]+)_([^_]+)\.nc$"


def parsear_nombre(nombre: str) -> Dict[str, Any]:
    m = re.match(PATRON_NOMBRE, nombre)
    if m:
        return {
            "variable_id": m.group(1), "table_id": m.group(2),
            "source_id":   m.group(3), "experiment_id": m.group(4),
            "member_id":   m.group(5), "grid_label": m.group(6),
            "time_range":  m.group(7), "filename": nombre,
        }
    return {"filename": nombre, "parse_error": True}


def extraer_anio(member_id: str) -> Optional[int]:
    m = re.search(r's(\d{4})', member_id)
    return int(m.group(1)) if m else None


def extraer_variant(member_id: str) -> Optional[str]:
    m = re.search(r'-(r\d+i\d+p\d+f\d+)', member_id)
    return m.group(1) if m else None


def cargar_metadatos(archivos_nc: List[Path]) -> pd.DataFrame:
    filas = []
    for f in tqdm(archivos_nc, desc="Leyendo metadatos"):
        info = parsear_nombre(f.name)
        info["file_size_mb"] = f.stat().st_size / (1024 * 1024)
        info["file_path"] = str(f)
        if "member_id" in info:
            info["init_year"]     = extraer_anio(info["member_id"])
            info["variant_label"] = extraer_variant(info["member_id"])
        filas.append(info)
    return pd.DataFrame(filas)


def calcular_completitud(df: pd.DataFrame) -> Dict[str, Dict]:
    completitud: Dict[str, Dict] = {}
    df = df[~df.get('parse_error', pd.Series(False, index=df.index)).astype(bool)]

    for (source, variant), grupo in df.groupby(['source_id', 'variant_label']):
        anios = sorted(int(a) for a in grupo['init_year'].dropna().unique())
        faltantes = sorted(set(range(min(anios), max(anios) + 1)) - set(anios))
        completitud[f"{source}_{variant}"] = {
            "total_files":      len(grupo),
            "init_years":       anios,
            "missing_years":    faltantes,
            "expected_size_mb": grupo['file_size_mb'].mean(),
        }
    return completitud


def generar_excel(df: pd.DataFrame, completitud: Dict[str, Dict],
                  ruta_salida: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Ejecutivo"

    # Titulo
    ws['A1'] = "INFORME DE CALIDAD DE DATOS NetCDF"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    fila = 3

    # --- Estadisticas generales ---
    ws[f'A{fila}'] = "ESTADISTICAS GENERALES"
    ws[f'A{fila}'].font = Font(bold=True, size=12)
    fila += 1
    ws[f'A{fila}'] = "Total de archivos:"
    ws[f'B{fila}'] = len(df)
    fila += 1
    ws[f'A{fila}'] = "Tamano total (MB):"
    ws[f'B{fila}'] = round(df['file_size_mb'].sum(), 2)
    fila += 1
    ws[f'A{fila}'] = "Combinaciones source_id/variant_label:"
    ws[f'B{fila}'] = len(completitud)
    fila += 2

    # --- Cobertura temporal ---
    ws[f'A{fila}'] = "COBERTURA TEMPORAL"
    ws[f'A{fila}'].font = Font(bold=True, size=12)
    fila += 1
    if 'init_year' in df.columns:
        anios = df['init_year'].dropna()
        if len(anios) > 0:
            ws[f'A{fila}'] = "Anios de inicializacion:"
            ws[f'B{fila}'] = f"{int(anios.min())} - {int(anios.max())}"
            fila += 1
            ws[f'A{fila}'] = "Total de anios unicos:"
            ws[f'B{fila}'] = len(anios.unique())
            fila += 1
    fila += 1

    # --- Calidad de datos ---
    ws[f'A{fila}'] = "CALIDAD DE DATOS"
    ws[f'A{fila}'].font = Font(bold=True, size=12)
    fila += 1
    total_faltantes = sum(len(v['missing_years']) for v in completitud.values())
    ws[f'A{fila}'] = "Anios faltantes:"
    ws[f'B{fila}'] = total_faltantes
    ws[f'C{fila}'] = "[OK] Completo" if total_faltantes == 0 else "[!] Incompleto"
    ws[f'C{fila}'].font = Font(color="00B050" if total_faltantes == 0 else "FF0000")
    fila += 2

    # --- Detalles por combinacion ---
    ws[f'A{fila}'] = "DETALLES POR COMBINACION"
    ws[f'A{fila}'].font = Font(bold=True, size=12)
    fila += 1
    for nombre, info in completitud.items():
        ws[f'A{fila}'] = nombre
        ws[f'A{fila}'].font = Font(bold=True)
        fila += 1
        ws[f'A{fila}'] = "  Archivos:"
        ws[f'B{fila}'] = info['total_files']
        fila += 1
        ws[f'A{fila}'] = "  Tamano esperado (MB):"
        ws[f'B{fila}'] = round(info['expected_size_mb'], 2)
        fila += 1
        if info['missing_years']:
            ws[f'A{fila}'] = "  Anios faltantes:"
            ws[f'B{fila}'] = ', '.join(map(str, info['missing_years']))
            fila += 1
        fila += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    wb.save(ruta_salida)
    print(f"[DONE] Guardado en: {ruta_salida}")


def main() -> None:
    directorio = Path(".")
    archivos_nc = list(directorio.rglob('*.nc'))

    if not archivos_nc:
        print("[X] No se encontraron archivos .nc")
        return

    print(f"Archivos encontrados: {len(archivos_nc)}")

    df = cargar_metadatos(archivos_nc)
    completitud = calcular_completitud(df)
    generar_excel(df, completitud, directorio / "resumen_ejecutivo.xlsx")


if __name__ == "__main__":
    main()
